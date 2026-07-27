import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

CURRENCY_SYMBOLS = {
    'USD':'$', 'EUR':'€', 'GBP':'£', 'JPY':'¥', 'INR':'₹',
    'CNY':'¥', 'CAD':'CA$', 'AUD':'A$', 'CHF':'CHF',
    'BRL':'R$', 'ZAR':'R', 'KRW':'₩', 'SGD':'S$', 'MXN':'$', 'NOK':'kr'
}

def format_currency(value, currency_code):
    sym = CURRENCY_SYMBOLS.get(currency_code, '$')
    if abs(value) >= 1e9:
        return f"{sym}{value/1e9:.2f} B"
    elif abs(value) >= 1e6:
        return f"{sym}{value/1e6:.2f} M"
    elif abs(value) >= 1e3:
        return f"{sym}{value/1e3:.2f} K"
    return f"{sym}{value:,.0f}"

def generate_pdf_report(result):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CustomTitle', parent=styles['Heading1'],
                              fontSize=24, alignment=TA_CENTER, spaceAfter=20))
    styles.add(ParagraphStyle(name='CustomHeading', parent=styles['Heading2'],
                              fontSize=16, spaceAfter=12, textColor=colors.HexColor('#34d399')))
    styles.add(ParagraphStyle(name='CustomBody', parent=styles['Normal'],
                              fontSize=11, spaceAfter=6))
    styles.add(ParagraphStyle(name='CustomCenter', parent=styles['Normal'],
                              fontSize=11, alignment=TA_CENTER, spaceAfter=6))

    elements = []

    elements.append(Paragraph("Gross Ecosystem Product (GEP) Report", styles['CustomTitle']))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['CustomCenter']))
    elements.append(Spacer(1, 0.5*cm))

    elements.append(Paragraph("1. Summary", styles['CustomHeading']))
    region = result.get('region', 'Unknown Region')
    currency = result.get('currency', 'INR')
    gep = result.get('gep', 0)
    epv = result.get('epv', 0)
    erv = result.get('erv', 0)
    ecv = result.get('ecv', 0)
    fv = result.get('fv', 0)

    data = [
        ["Region", region],
        ["Currency", currency],
        ["Total GEP", format_currency(gep, currency)],
        ["Provisioning (EPV)", format_currency(epv, currency)],
        ["Regulating (ERV)", format_currency(erv, currency)],
        ["Cultural (ECV)", format_currency(ecv, currency)],
        ["Fauna (FV)", format_currency(fv, currency)],
    ]
    table = Table(data, colWidths=[4*cm, 8*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#1a2a20')),
        ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#d1fae5')),
        ('BACKGROUND', (1,0), (1,-1), colors.HexColor('#0d1c14')),
        ('TEXTCOLOR', (1,0), (1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#2a4a3a')),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))

    elements.append(Paragraph("2. Service Composition", styles['CustomHeading']))
    total = gep if gep > 0 else 1
    components = result.get('components', [])
    if not components:
        components = [
            {'label': 'Provisioning', 'value': epv, 'color': '#34d399'},
            {'label': 'Regulating', 'value': erv, 'color': '#60a5fa'},
            {'label': 'Cultural', 'value': ecv, 'color': '#fbbf24'},
            {'label': 'Fauna', 'value': fv, 'color': '#f472b6'}
        ]

    max_val = max([c['value'] for c in components if c['value'] > 0], default=1)
    chart_width = 12 * cm
    chart_height = 1.2 * cm * len(components)

    d = Drawing(chart_width, chart_height + 1*cm)
    y_pos = chart_height
    for comp in components:
        val = comp['value']
        if val == 0:
            continue
        pct = (val / total) * 100
        bar_width = (val / max_val) * chart_width * 0.85
        d.add(String(0, y_pos - 0.4*cm, comp['label'], fontSize=10))
        d.add(Rect(4*cm, y_pos - 0.6*cm, bar_width, 0.6*cm,
                   fillColor=colors.HexColor(comp['color']), strokeColor=None))
        d.add(String(4*cm + bar_width + 0.2*cm, y_pos - 0.4*cm,
                     f"{pct:.1f}%", fontSize=10))
        y_pos -= 1.2 * cm

    elements.append(d)
    elements.append(Spacer(1, 0.5*cm))

    elements.append(Paragraph("3. Input Parameters", styles['CustomHeading']))
    inputs_data = []
    if 'inputs' in result:
        for k, v in result['inputs'].items():
            if v and float(v) > 0:
                label = k.replace('p-', '').replace('r-', '').replace('c-', '').replace('fa-', '').replace('f-', '').replace('-', ' ').title()
                inputs_data.append([label, str(v)])
    else:
        inputs_data = [
            ["Region", region],
            ["Area (km²)", result.get('area', 'N/A')],
            ["Population", result.get('population', 'N/A')],
        ]

    if inputs_data:
        t = Table(inputs_data, colWidths=[5*cm, 7*cm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#1a2a20')),
            ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#d1fae5')),
            ('BACKGROUND', (1,0), (1,-1), colors.HexColor('#0d1c14')),
            ('TEXTCOLOR', (1,0), (1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#2a4a3a')),
            ('ALIGN', (1,0), (1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(t)

    elements.append(Spacer(1, 1*cm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#34d399')))
    elements.append(Paragraph("Generated by GEP Calculator — Ecological Accounting & Nature Intelligence",
                              ParagraphStyle('Footer', parent=styles['Normal'],
                                             fontSize=9, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(result):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a2a20", end_color="1a2a20", fill_type="solid")

    summary_data = [
        ["Metric", "Value"],
        ["Region", result.get('region', 'Unknown')],
        ["Currency", result.get('currency', 'INR')],
        ["Total GEP", result.get('gep', 0)],
        ["Provisioning (EPV)", result.get('epv', 0)],
        ["Regulating (ERV)", result.get('erv', 0)],
        ["Cultural (ECV)", result.get('ecv', 0)],
        ["Fauna (FV)", result.get('fv', 0)],
        ["Population", result.get('population', 'N/A')],
        ["Area (km²)", result.get('area', 'N/A')],
    ]
    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                if col_idx == 2 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet("Inputs")
    ws2.append(["Parameter", "Value"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    if 'inputs' in result:
        for k, v in result['inputs'].items():
            if v:
                label = k.replace('p-', '').replace('r-', '').replace('c-', '').replace('fa-', '').replace('f-', '').replace('-', ' ').title()
                ws2.append([label, v])
    else:
        ws2.append(["No detailed inputs captured", ""])

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()